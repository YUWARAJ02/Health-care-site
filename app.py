import numpy as np
import pandas as pd
from flask import Flask, request, render_template, jsonify, send_file, session, redirect, url_for, make_response
import pickle as p
import datetime
app=Flask(__name__)
# app.secret_key = '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'
chronic_model=p.load(open('CKD.pk1','rb'))
heart_model=p.load(open('HDP.pk1','rb'))
cd_model=p.load(open('CD.pk1','rb'))
## ---------------------loading csv file -----------------------
import openpyxl
##-----------------------------------------------------------------

today = datetime.date.today()
date=today.strftime('%d/%m/%Y')

@app.route('/')
def HOME():
    return render_template('final_home.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/feedback')
def feedback():
    return render_template('feedback.html')
##  -------------------HEART-------------------------------------
@app.route('/heart_prediction')
def heart():
    return render_template('prediction_heart.html')

@app.route('/common')
def common():
    return render_template('common.html')

@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/download_file')
def download_file():
    filename = request.args.get('filename')
    return send_file(filename, as_attachment=True)



@app.route('/submit-checkbox-values', methods=['POST'])
def submit_checkbox_values():
    data = request.get_json()
    
    my_array = np.array(list(data.values()))
    my_data = np.array(list(data.keys()))
    print(my_array)
    print(my_data)
    df = pd.DataFrame(data=[my_array], columns=my_data)
    print('shape : ' + str(df.shape))
    df['fluid_overload.1'] = 0
    output=cd_model.predict(df)
    print(output)
    response = {'message': str(output[0])}
    return jsonify(response)


@app.route('/heart_predict',methods=['POST'])
def h_prediction():    
    wb = openpyxl.load_workbook('heart_data.xlsx')
    heart_data=wb['Sheet1']
    next_row = heart_data.max_row + 1
    heart_data.cell(row=next_row, column=1, value=date)
    form_value=request.form.values()    
    data=[]
    for x in form_value:
        data.append(pd.to_numeric(x).astype(float))    
    for i in range(len(data)):
        heart_data.cell(row=next_row, column=i+2, value=data[i])      
    
    features_value=[np.array(data)]  
    features_name=['age','sex','cp','trestbps',
                   'chol','fbs','thalach','oldpeak','slope','ca','thal']
    df=pd.DataFrame(features_value, columns=features_name)    
    output=heart_model.predict(df)   
    
    if(output==1):
        heart_data.cell(row=next_row, column=13, value='affected')
        wb.save('heart_data.xlsx')
        return render_template('prediction_heart.html' , pred=' You are affected Heart Disease. So, please concern a Doctor')
    else:
        heart_data.cell(row=next_row, column=13, value='not affected')
        wb.save('heart_data.xlsx')
        return render_template('prediction_heart.html' , pred='you are not affected by Heart Disease')
##---------------------------------------------------------------


##-------------------KIDNEY--------------------------------------
@app.route('/kidney_prediction')
def kidney():
    return render_template('chronic_kidney.html')

@app.route('/chronic_predict',methods=['POST'])
def c_prediction():
    wb = openpyxl.load_workbook('kidney_data.xlsx')
    kidney_data=wb['Sheet1']
    next_row = kidney_data.max_row + 1  
    kidney_data.cell(row=next_row, column=1, value=date)  
    form_value=request.form.values()    
    data=[]
    for x in form_value:
        data.append(pd.to_numeric(x).astype(float))
    for i in range(len(data)):
        kidney_data.cell(row=next_row, column=i+2, value=data[i])    
    features_value=[np.array(data)]  
    features_name=['age','blood_urea','blood glucose random','coronary_artery_disease',
                   'anemia','pus_cell','red_blood_cells','diabetesmellitus','pedal_edema']
    df=pd.DataFrame(features_value, columns=features_name)
    
    output=chronic_model.predict(df)
    if(output==0):
        kidney_data.cell(row=next_row, column=11, value='ckd')
        wb.save('kidney_data.xlsx')
        return render_template('chronic_kidney.html' , pred='Oops!! You have Kidney Chronic Disease. So, please concern a Doctor')
    else:
        kidney_data.cell(row=next_row, column=11, value='notckd')
        wb.save('kidney_data.xlsx')
        return render_template('chronic_kidney.html' , pred='you are not affected by Chronic kidney Disease')


@app.route('/validate',methods=['POST'])
def validate():
    username = request.json["username"]
    password = request.json["password"]
    
    if username == "admin" and password == "admin":
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "message": "Incorrect username or password"})




if __name__=='__main__':
    app.run(debug=True)